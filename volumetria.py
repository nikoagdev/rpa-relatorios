import os
from datetime import date, timedelta
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --- CONFIGURAÇÕES DINÂMICAS ---
# Pega o caminho absoluto da pasta onde este script está localizado
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Define os caminhos relativos à localização do script
PASTA_RELATORIOS = os.path.join(BASE_DIR, 'relatorios_convertidos') # <-- MUDANÇA AQUI
ARQUIVO_SAIDA = os.path.join(BASE_DIR, 'Volumetria_Final.xlsx')    # <-- MUDANÇA AQUI


# --- O RESTANTE DO SEU CÓDIGO PERMANECE O MESMO ---
# Nenhuma mudança é necessária na lógica de cálculo ou formatação.
# Cole o restante do seu código original de volumetria.py aqui, a partir de:
# COLUNA_DATA = 'Data Solicitação'
# ... até o final ...

COLUNA_DATA = 'Data Solicitação'
COLUNA_CANAL = 'Área'
COLUNA_SOLICITACAO = 'Solicitação'
COLUNA_STATUS = 'Status'
SLA_DIAS_UTEIS_PADRAO = 2
SLA_DIAS_UTEIS_SPO = 3
FERIADOS = [date(2025, 1, 1), date(2025, 3, 3), date(2025, 3, 4), date(2025, 4, 18), date(2025, 4, 21), date(2025, 5, 1), date(2025, 6, 19), date(2025, 9, 7), date(2025, 10, 12), date(2025, 11, 2), date(2025, 11, 15), date(2025, 12, 25)]
AREAS_DASHBOARD = ['COORDENAÇÃO DE CURSO', 'CRA PEDAGÓGICO', 'SECRETARIA ACADÊMICA - SERVIÇOS ACADÊMICOS - ATENDIMENTO N1', 'EAD POS - ATENDIMENTO', 'EAD PÓS - SECRETARIA ACADÊMICA - SERVIÇOS ACADÊMICOS - ATENDIMENTO N1', 'EAD PROF - ATENDIMENTO', 'EAD PROF - SECRETARIA ACADÊMICA - SERVIÇOS ACADÊMICOS - ATENDIMENTO N1', 'EAD TEC - ATENDIMENTO', 'EAD TEC - CRA PEDAGÓGICO', 'EAD TEC - SECRETARIA ACADÊMICA - SERVIÇOS ACADÊMICOS - ATENDIMENTO N1', 'FINANCEIRO - ATENDIMENTO N1', 'EAD POS - FINANCEIRO - ATENDIMENTO N1', 'EAD PROF - FINANCEIRO - ATENDIMENTO N1', 'EAD TEC - FINANCEIRO - ATENDIMENTO N1', 'LOGÍSTICA - ATENDIMENTO N1', 'DECLARAÇÃO DE VÍNCULO - SPO']
AREAS_PARA_DISTRIBUIR = ['COORDENAÇÃO DE CURSO', 'CRA PEDAGÓGICO', 'LOGÍSTICA - ATENDIMENTO N1', 'FINANCEIRO - ATENDIMENTO N1', 'SECRETARIA ACADÊMICA - SERVIÇOS ACADÊMICOS - ATENDIMENTO N1', 'EAD POS - FINANCEIRO - ATENDIMENTO N1', 'EAD PROF - FINANCEIRO - ATENDIMENTO N1', 'EAD TEC - FINANCEIRO - ATENDIMENTO N1', 'EAD POS - ATENDIMENTO', 'EAD PÓS - SECRETARIA ACADÊMICA - SERVIÇOS ACADÊMICOS - ATENDIMENTO N1', 'EAD PROF - ATENDIMENTO', 'EAD PROF - SECRETARIA ACADÊMICA - SERVIÇOS ACADÊMICOS - ATENDIMENTO N1', 'EAD TEC - ATENDIMENTO', 'EAD TEC - SECRETARIA ACADÊMICA - SERVIÇOS ACADÊMICOS - ATENDIMENTO N1', 'EAD TEC - CRA PEDAGÓGICO']


def calcular_vencimento(data_solicitacao, sla_dias):
    data_vencimento = data_solicitacao
    dias_adicionados = 0
    while dias_adicionados < sla_dias:
        data_vencimento += timedelta(days=1)
        if data_vencimento.weekday() < 5 and data_vencimento not in FERIADOS:
            dias_adicionados += 1
    return data_vencimento

def calcular_dias_uteis_entre(d1, d2):
    if d2 < d1: return -1
    dias_uteis = 0; d_temp = d1
    while d_temp < d2:
        d_temp += timedelta(days=1)
        if d_temp.weekday() < 5 and d_temp not in FERIADOS: dias_uteis += 1
    return dias_uteis

def apply_external_border(ws, cell_range):
    thin = Side(style='thin')
    rows = list(ws[cell_range])
    for r_idx, row in enumerate(rows):
        for c_idx, cell in enumerate(row):
            current_border = cell.border.copy()
            if r_idx == 0: current_border.top = thin
            if r_idx == len(rows) - 1: current_border.bottom = thin
            if c_idx == 0: current_border.left = thin
            if c_idx == len(row) - 1: current_border.right = thin
            cell.border = current_border

def main():
    print("\n--- INICIANDO SCRIPT DE VOLUMETRIA ---")
    try:
        arquivos_excel = [f for f in os.listdir(PASTA_RELATORIOS) if f.endswith('.xlsx') and not f.startswith('~$')]
    except FileNotFoundError: return print(f"ERRO: A pasta '{PASTA_RELATORIOS}' não foi encontrada.")
    if not arquivos_excel: return print(f"Nenhum arquivo .xlsx encontrado em '{PASTA_RELATORIOS}'.")

    lista_dfs_padronizados = []
    for arquivo in arquivos_excel:
        caminho_arquivo = os.path.join(PASTA_RELATORIOS, arquivo)
        try:
            df_lido = pd.read_excel(caminho_arquivo)
            df_temp = pd.DataFrame()
            if 'spo' in arquivo.lower():
                df_temp[COLUNA_SOLICITACAO] = df_lido['Protocolo']
                df_temp[COLUNA_CANAL] = 'DECLARAÇÃO DE VÍNCULO - SPO'
                df_temp[COLUNA_STATUS] = df_lido['Status Protocolo']
                df_temp[COLUNA_DATA] = df_lido['Data Prot.']
            else:
                df_temp = df_lido
            colunas_necessarias = [COLUNA_SOLICITACAO, COLUNA_CANAL, COLUNA_STATUS, COLUNA_DATA]
            if all(c in df_temp.columns for c in colunas_necessarias):
                lista_dfs_padronizados.append(df_temp[colunas_necessarias])
        except Exception as e:
            print(f" !!! Erro ao processar o arquivo {arquivo}: {e}")

    if not lista_dfs_padronizados: return print("Nenhum dado válido foi lido. Encerrando.")
    df_completo = pd.concat(lista_dfs_padronizados, ignore_index=True)
    df_completo.dropna(subset=[COLUNA_CANAL, COLUNA_STATUS, COLUNA_DATA], inplace=True)
    df_completo[COLUNA_STATUS] = df_completo[COLUNA_STATUS].astype(str).str.strip()
    
    print("\nAplicando regras de filtro por status e área...")
    cond_coord = (df_completo[COLUNA_CANAL].str.contains('COORDENAÇÃO', na=False)) & (df_completo[COLUNA_STATUS].str.lower() == 'nova')
    cond_spo = (df_completo[COLUNA_CANAL] == 'DECLARAÇÃO DE VÍNCULO - SPO') & (df_completo[COLUNA_STATUS].str.lower() == 'protocolo em andamento')
    cond_sae = (~df_completo[COLUNA_CANAL].str.contains('COORDENAÇÃO', na=False)) & (df_completo[COLUNA_CANAL] != 'DECLARAÇÃO DE VÍNCULO - SPO') & (df_completo[COLUNA_STATUS].str.lower() == 'encaminhada')
    
    df_consolidado = df_completo[cond_coord | cond_spo | cond_sae].copy()
    print(f"Filtrado para {len(df_consolidado)} solicitações válidas para cálculo de SLA.")
    
    df_volumetria = pd.DataFrame()
    if not df_consolidado.empty:
        df_consolidado[COLUNA_DATA] = pd.to_datetime(df_consolidado[COLUNA_DATA], dayfirst=True, errors='coerce').dt.date
        df_consolidado.dropna(subset=[COLUNA_DATA, COLUNA_SOLICITACAO], inplace=True)
        hoje = date.today()
        volumetria = {}
        for _, linha in df_consolidado.iterrows():
            canal = str(linha[COLUNA_CANAL]).strip()
            data_solicitacao = linha[COLUNA_DATA]
            if pd.isna(data_solicitacao): continue
            if canal not in volumetria:
                volumetria[canal] = {'ATRASOS': 0, 'HOJE': 0, 'D+1': 0, 'D+2': 0, 'D+3': 0}
            sla_aplicado = SLA_DIAS_UTEIS_SPO if canal == 'DECLARAÇÃO DE VÍNCULO - SPO' else SLA_DIAS_UTEIS_PADRAO
            data_vencimento = calcular_vencimento(data_solicitacao, sla_aplicado)
            categoria = ''
            if hoje > data_vencimento: categoria = 'ATRASOS'
            else:
                diff = calcular_dias_uteis_entre(hoje, data_vencimento)
                if diff == 0: categoria = 'HOJE'
                elif diff == 1: categoria = 'D+1'
                elif diff == 2: categoria = 'D+2'
                elif diff == 3: categoria = 'D+3'
            if categoria: volumetria[canal][categoria] += 1
        df_volumetria = pd.DataFrame.from_dict(volumetria, orient='index')

    colunas_ordenadas = ['ATRASOS', 'HOJE', 'D+1', 'D+2', 'D+3']
    df_volumetria = df_volumetria.reindex(AREAS_DASHBOARD).reindex(columns=colunas_ordenadas).fillna(0).astype(int)

    try:
        if os.path.exists(ARQUIVO_SAIDA):
            print(f"Removendo versão anterior de '{os.path.basename(ARQUIVO_SAIDA)}'...")
            try:
                os.remove(ARQUIVO_SAIDA)
            except PermissionError:
                print(f" !!! AVISO: Não foi possível remover o arquivo. Ele pode estar aberto em outro programa.")
                return

        print(f"Gerando novo arquivo de volumetria em '{ARQUIVO_SAIDA}'...")
        with pd.ExcelWriter(ARQUIVO_SAIDA, engine='openpyxl') as writer:
            df_volumetria.to_excel(writer, sheet_name='Dashboard Volumetria', index_label='ÁREA')
            df_para_tarefas = df_consolidado[df_consolidado[COLUNA_CANAL].isin(AREAS_PARA_DISTRIBUIR)].copy()
            df_tarefas_priorizadas = df_para_tarefas.sort_values(by=COLUNA_DATA, ascending=True)
            df_tarefas_priorizadas[[COLUNA_SOLICITACAO, COLUNA_DATA, COLUNA_CANAL]].to_excel(writer, sheet_name='Tarefas Priorizadas', index=False)

            print("\nGerando abas de distribuição por área...")
            for area in AREAS_PARA_DISTRIBUIR:
                df_area = df_consolidado[df_consolidado[COLUNA_CANAL] == area].copy()
                if not df_area.empty:
                    df_area_sorted = df_area.sort_values(by=COLUNA_DATA, ascending=True)
                    nome_da_aba = area.replace('/', '-').replace('\\', '-').replace(':', '-')[:31]
                    df_area_sorted[[COLUNA_SOLICITACAO, COLUNA_DATA]].to_excel(writer, sheet_name=nome_da_aba, index=False)

            print("\nAplicando formatação visual final...")
            ws = writer.sheets['Dashboard Volumetria']
            
            ws.sheet_view.showGridLines = False
            header_font = Font(bold=True, color="FFFFFF", name='Calibri', size=11)
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            center_align = Alignment(horizontal='center', vertical='center')
            left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
            bold_font = Font(bold=True, name='Calibri', size=11)
            red_font = Font(color="C00000", bold=True, name='Calibri', size=11)
            total_row_idx = len(AREAS_DASHBOARD) + 2
            ws[f'A{total_row_idx}'] = "TOTAL"
            
            for col_idx in range(1, 7):
                col_letter = get_column_letter(col_idx)
                for row_idx in [1, total_row_idx]:
                    cell = ws[f'{col_letter}{row_idx}']
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                if 1 < col_idx < 7:
                    ws[f'{col_letter}{total_row_idx}'].value = f'=SUM({col_letter}2:{col_letter}{total_row_idx - 1})'
            
            for row_num in range(2, total_row_idx):
                ws[f'A{row_num}'].alignment = left_align
                ws[f'A{row_num}'].font = bold_font
                for col_num in range(2, 7):
                    ws.cell(row=row_num, column=col_num).alignment = center_align
                cell_atraso = ws.cell(row=row_num, column=2)
                if isinstance(cell_atraso.value, (int, float)) and cell_atraso.value >= 0:
                    cell_atraso.font = red_font
                
                area_name = ws.cell(row=row_num, column=1).value
                if area_name != 'DECLARAÇÃO DE VÍNCULO - SPO':
                    cell_d3 = ws.cell(row=row_num, column=6)
                    if cell_d3.value != 0:
                        cell_d3.value = 0 
            
            apply_external_border(ws, "A1:F1")
            apply_external_border(ws, f"A{total_row_idx}:F{total_row_idx}")
            
            start_row_t1 = total_row_idx + 2
            ws[f'B{start_row_t1}'] = 'ATIVIDADE COMPLEMENTAR'
            ws[f'C{start_row_t1}'] = 'ATRASOS'
            ws[f'D{start_row_t1}'] = 'NO PRAZO'
            ws[f'B{start_row_t1+1}'] = 'EXTERNAS'
            ws[f'B{start_row_t1+2}'] = 'INTERNAS'
            for row in ws[f"B{start_row_t1}:D{start_row_t1+2}"]:
                for cell in row:
                    cell.alignment = center_align
                    if cell.coordinate in [f'B{start_row_t1}', f'C{start_row_t1}', f'D{start_row_t1}', f'B{start_row_t1+1}', f'B{start_row_t1+2}']:
                        cell.font = header_font
                        cell.fill = header_fill
            
            apply_external_border(ws, f"B{start_row_t1}:D{start_row_t1+2}")
            
            start_row_t2 = start_row_t1 + 4
            ws[f'B{start_row_t2}'] = 'TOTAL GERAL EM ATRASO'
            ws[f'C{start_row_t2}'] = f'=SUM(B{total_row_idx},C{start_row_t1+1}:C{start_row_t1+2})'
            ws[f'D{start_row_t2}'] = 'TOTAL NO PRAZO'
            ws[f'E{start_row_t2}'] = f'=SUM(C{total_row_idx}:F{total_row_idx},D{start_row_t1+1}:D{start_row_t1+2})'
            for cell_coord_str in [f'B{start_row_t2}', f'C{start_row_t2}', f'D{start_row_t2}', f'E{start_row_t2}']:
                cell = ws[cell_coord_str]
                cell.font = bold_font
                cell.alignment = center_align
            ws[f'B{start_row_t2}'].fill = header_fill
            ws[f'D{start_row_t2}'].fill = header_fill
            ws[f'B{start_row_t2}'].font = header_font
            ws[f'D{start_row_t2}'].font = header_font
            
            apply_external_border(ws, f"B{start_row_t2}:E{start_row_t2}")
            
            ws.column_dimensions['A'].width = 70
            for col_letter in ['B', 'C', 'D', 'E', 'F']:
                ws.column_dimensions[col_letter].width = 25

        print("-" * 30)
        print(f"Sucesso! Arquivo '{os.path.basename(ARQUIVO_SAIDA)}' foi gerado com todas as abas.")
        print("-" * 30)

        print("\nAbrindo o arquivo de volumetria final...")
        os.startfile(ARQUIVO_SAIDA)

    except Exception as e:
        print(f"ERRO INESPERADO ao salvar, formatar ou abrir o arquivo: {e}")

if __name__ == '__main__':
    main()